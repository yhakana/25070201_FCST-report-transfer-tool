from openpyxl import load_workbook
from datetime import datetime
import calendar
import re

def gen_next_6months_titles():
    months = []
    now = datetime.now()
    for i in range(6):
        y = now.year
        m = now.month + i
        if m > 12:
            y += (m - 1) // 12
            m = (m - 1) % 12 + 1
        if i == 0 and now.month == 6 and now.year == 2025:  # 若起始為 Jun 2025
            months.append("Jun'25")
        else:
            # 這裡預設 Jul, Aug ... 若你有特殊命名再加條件
            if y == 2025 and m == 6:
                months.append("Jun'25")
            elif y == 2025 and m == 7:
                months.append("Jul")
            elif y == 2025 and m == 8:
                months.append("Aug")
            elif y == 2025 and m == 9:
                months.append("Sep")
            elif y == 2025 and m == 10:
                months.append("Oct")
            elif y == 2025 and m == 11:
                months.append("Nov")
            elif y == 2025 and m == 12:
                months.append("Dec")
            elif y == 2026 and m == 1:
                months.append("2026-Jan")
            elif y == 2026 and m == 2:
                months.append("Feb")
            elif y == 2026 and m == 3:
                months.append("Mar")
            elif y == 2026 and m == 4:
                months.append("Apr")
            elif y == 2026 and m == 5:
                months.append("May")
            else:
                months.append(f"{calendar.month_abbr[m]}'{str(y)[-2:]}")
    return months

def find_multilevel_header(ws, main_titles, month_titles, search_limit=30):
    for r in range(1, search_limit):
        row1 = [str(cell.value).strip() if cell.value else "" for cell in ws[r]]
        main_hits = [h for h in main_titles if h in row1]
        if len(main_hits) >= 2:
            row2 = [str(cell.value).strip() if cell.value else "" for cell in ws[r+1]]
            month_hits = [m for m in month_titles if m in row2]
            if len(month_hits) >= 2:
                return r, r+1
    raise ValueError("找不到雙列標題")

def merge_multilevel_header(header1, header2):
    """
    主標題遇到空白會自動向左補值，產生最終欄位名稱
    - header1: 主標題（list）
    - header2: 副標題（list）
    """
    header1_filled = header1.copy()
    last_main = ""
    for i in range(len(header1_filled)):
        if header1_filled[i]:
            last_main = header1_filled[i]
        else:
            header1_filled[i] = last_main

    final_headers = []
    for h1, h2 in zip(header1_filled, header2):
        h1 = h1.strip() if isinstance(h1, str) else ""
        h2 = h2.strip() if isinstance(h2, str) else ""
        if h1 and h2:
            final_headers.append(f"{h1}_{h2}")
        elif not h1 and h2:
            final_headers.append(h2)
        elif h1 and not h2:
            final_headers.append(h1)
        else:
            final_headers.append("")
    return final_headers

def make_headers_unique(headers):
    counts = {}
    result = []
    for h in headers:
        if h == "":
            result.append("")
            continue
        if h not in counts:
            counts[h] = 1
            result.append(h)
        else:
            counts[h] += 1
            result.append(f"{h}_{counts[h]}")
    return result

def extract_group_month_fields(headers):
    # 輸出 dict: {主題: [(月份, 欄位名)]}
    group_fields = {}
    pattern = r"^(.+)_((?:Jan|Feb|Mar|Apr|May|Jun'25|Jun|Jul|Aug|Sep|Oct|Nov|Dec)(?:'\d\d|\-\d\d\d\d)?)(?:_(\d+))?$"
    for h in headers:
        m = re.match(pattern, h)
        if m:
            group = m.group(1)
            month = m.group(2)
            if group not in group_fields:
                group_fields[group] = []
            group_fields[group].append((month, h))
    return group_fields

def next_n_month_names(n, header_months):
    now = datetime.now()
    y, m = now.year, now.month
    months = []
    cnt = 0
    all_header_months = set(header_months)
    while cnt < n:
        # 以欄位實際出現的名稱為準，自動對應年份與特殊寫法
        # 嘗試 "Jul"、"Jul'25"、"2026-Jan" 這種格式
        this_year = y
        m_str_std = datetime(this_year, m, 1).strftime("%b")      # "Jul"
        m_str_jun25 = "Jun'25" if m == 6 and "Jun'25" in all_header_months else None
        m_str_year = f"{this_year}-{m_str_std}" if f"{this_year}-{m_str_std}" in all_header_months else None
        # 優先順序: Jun'25 > yyyy-Mmm > Mmm
        if m_str_jun25:
            m_str = m_str_jun25
        elif m_str_year:
            m_str = m_str_year
        else:
            m_str = m_str_std
        months.append(m_str)
        m += 1
        if m > 12:
            m = 1
            y += 1
        cnt += 1
    return months

if __name__ == "__main__":
    # 1. 載入檔案
    source = "250702162359805.xlsm"
    wb = load_workbook(source, keep_vba=True)
    ws = wb["CP Summary"]  # 你要抓 FT Summary 也只要 ws = wb["FT Summary"]

    # 你的設定
    main_titles = ['Machine O/H (set)', 'Machine require(set/Wk)', 'Idling tester (set)']
    months = gen_next_6months_titles()
    
    header_row1, header_row2 = find_multilevel_header(ws, main_titles + ["Process", "Tester", "Customer"], months)

    header1 = [str(cell.value).strip() if cell.value else "" for cell in ws[header_row1]]
    header2 = [str(cell.value).strip() if cell.value else "" for cell in ws[header_row2]]

    # 組合欄位名稱
    final_headers = merge_multilevel_header(header1, header2)
    unique_headers = make_headers_unique(final_headers)

    # 取得所有主題_月份型態的欄位
    group_month_fields = extract_group_month_fields(unique_headers)
    # 取得所有月份名
    all_months = sorted({month for fields in group_month_fields.values() for (month, _) in fields})

    # print(group_month_fields.keys())

    # 取得當月起六個月的月份名（以所有 header 出現過的月份為依據）
    target_months = next_n_month_names(6, all_months)
    print("自動判斷本月起六個月：", target_months)

    # 為每個主題挑出連續六個月的欄位名
    target_fields_by_group = {}
    for group, fields in group_month_fields.items():
        # 這個主題下所有 (月份, 欄位名)
        month_to_field = {month: h for (month, h) in fields}
        # 依序取對應的欄位名（找不到用None填）
        target_fields = [month_to_field.get(month) for month in target_months]
        target_fields_by_group[group] = target_fields

    print("\n主題及自動對應六個月的欄位名：")
    for group, fields in target_fields_by_group.items():
        print(f"{group}: {fields}")

    fill_columns = ["Process", "Tester", "Customer"]
    last_values = {col: None for col in fill_columns}

    # 讀取資料列
    data_rows = []
    for row in ws.iter_rows(min_row=header_row2+1, values_only=True):
        if all(cell is None or cell == "" for cell in row):
            continue
        row_data = list(row)
        if len(row_data) < len(unique_headers):
            row_data += [None] * (len(unique_headers) - len(row_data))
        elif len(row_data) > len(unique_headers):
            row_data = row_data[:len(unique_headers)]
        row_dict = dict(zip(unique_headers, row_data))
        for col in fill_columns:
            if not row_dict.get(col):  # 若是 None 或 ""
                row_dict[col] = last_values[col]
            else:
                last_values[col] = row_dict[col]
        data_rows.append(row_dict)

    # 輸出每一列各主題下六個月的資料
    print("\n範例：前2筆資料（每個主題六個月）")
    for row in data_rows[:2]:
        print(f"Process: {row.get('Process', '')}, Tester: {row.get('Tester', '')}, Customer: {row.get('Customer', '')}")
        for group, fields in target_fields_by_group.items():
            values = [row.get(f) if f else None for f in fields]
            print(f"{group}: {values}")
    print("-" * 30)

    # # 資料讀取（從 header_row_2+1 開始）
    # data_rows = []
    # for row in ws.iter_rows(min_row=header_row2+1, values_only=True):
    #     # 檢查是不是空列
    #     if all(cell is None or cell == "" for cell in row):
    #         continue
    #     # 資料欄數與標題不符自動補 None
    #     row_data = list(row) + [None] * (len(unique_headers) - len(row))
    #     row_dict = dict(zip(unique_headers, row_data))
    #     data_rows.append(row_dict)

    # # 示範：取 Machine require(set/Wk) 下面六個月資料
    # target_main = "Machine require(set/Wk)"
    # target_months = ["Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]  # 根據你的需求調整
    # target_fields = [f"{target_main}_{m}" for m in target_months]

    # print("\n前3列 Machine require 六個月資料：")
    # for row in data_rows[:3]:
    #     print(f"Process: {row.get('Process', '')}, Tester: {row.get('Tester', '')}, Customer: {row.get('Customer', '')}")
    #     values = [row.get(field, "") for field in target_fields]
    #     print(values)

