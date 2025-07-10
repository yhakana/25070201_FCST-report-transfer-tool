from openpyxl import load_workbook

def find_multilevel_header(ws, key_headers, month_candidates, search_limit=30):
    """
    自動找出多層標題的起始row index（回傳: (row_idx1, row_idx2)）
    key_headers: 主要標題名清單（如 Process、Tester、Machine O/H...）
    month_candidates: 月份標題可能的字串list
    search_limit: 最多搜尋前幾列
    """
    for r in range(1, search_limit):
        row1 = [str(cell.value).strip() if cell.value else "" for cell in ws[r]]
        hits = [h for h in key_headers if h in row1]
        if len(hits) >= 2:  # 至少2個以上大標題
            row2 = [str(cell.value).strip() if cell.value else "" for cell in ws[r+1]]
            month_hits = [m for m in month_candidates if m in row2]
            if len(month_hits) >= 2:  # 至少2個月份
                return r, r+1
    raise ValueError("找不到雙列標題的位置")

def get_combined_headers(ws, row_idx1, row_idx2):
    """
    將兩列標題合併產生唯一欄位名稱
    """
    row1 = [str(cell.value).strip() if cell.value else "" for cell in ws[row_idx1]]
    row2 = [str(cell.value).strip() if cell.value else "" for cell in ws[row_idx2]]
    max_len = max(len(row1), len(row2))
    headers = []
    for i in range(max_len):
        h1 = row1[i] if i < len(row1) else ""
        h2 = row2[i] if i < len(row2) else ""
        if h1 and h2:
            headers.append(f"{h1} - {h2}")
        elif h1:
            headers.append(h1)
        elif h2:
            headers.append(h2)
        else:
            headers.append("")
    return headers

# === 主程式測試區段 ===

# 載入 Excel 檔案
source = "250702162359805.xlsm"
wb = load_workbook(source, keep_vba=True)
ws = wb["CP Summary"]  # 或 "FT Summary"，可自行切換

# 你預期的主標題和月份清單
key_headers = ['Process', 'Tester', 'Customer', 'Machine O/H (set)', 'Machine require(set/Wk)', "idling tester (set)"]
month_candidates = ["Jun'25", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec", "2026-Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul"]

# 自動搜尋標題列
row_idx1, row_idx2 = find_multilevel_header(ws, key_headers, month_candidates)
print(f"標題列位置：{row_idx1}、{row_idx2}")

# 組合標題欄位
combined_headers = get_combined_headers(ws, row_idx1, row_idx2)
print("合併後的標題名：")
for idx, h in enumerate(combined_headers):
    print(f"{idx+1}: {h}")

# 你之後可以根據 combined_headers 做欄位對應資料抓取
