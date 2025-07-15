from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import calendar
import re
import os
from collections import defaultdict
import logging

LOG_DIR = "log"
if not os.path.exists(LOG_DIR):
    os.makedirs(LOG_DIR)
today_str = datetime.now().strftime('%Y%m%d')
log_file = os.path.join(LOG_DIR, f"exec_{today_str}.log")

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(name)s: %(message)s",
    handlers=[
        logging.FileHandler(log_file, encoding="utf-8"),
        logging.StreamHandler()
    ]
)

# 欄位寬度與格式設定
col_widths = [16, 22, 22, 25, 12, 12, 12, 12, 12, 12, 14]
header_fill = PatternFill(start_color="FDE8D7", end_color="FDE8D7", fill_type="solid")
bold_font = Font(bold=True)
calibri_bold = Font(name='Calibri', bold=True)
thin = Side(border_style="thin", color="000000")
red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
border = Border(top=thin, left=thin, right=thin, bottom=thin)

def gen_next_6months_titles():
    try:
        months = []
        now = datetime.now()
        for i in range(6):
            y = now.year
            m = now.month + i
            if m > 12:
                y += (m - 1) // 12
                m = (m - 1) % 12 + 1
            if i == 0 and now.month == 6 and now.year == 2025:
                months.append("Jun'25")
            else:
                if y == 2025 and m == 6:
                    months.append("Jun'25")
                elif y == 2025 and m == 7:
                    months.append("Jul")
                elif y == 2025 and m == 8:
                    months.append("Aug")
                elif y == 2025 and m == 9:
                    months.append("Sep")
                elif y == 2025 and m == 10:
                    months.append("Oct")
                elif y == 2025 and m == 11:
                    months.append("Nov")
                elif y == 2025 and m == 12:
                    months.append("Dec")
                elif y == 2026 and m == 1:
                    months.append("2026-Jan")
                elif y == 2026 and m == 2:
                    months.append("Feb")
                elif y == 2026 and m == 3:
                    months.append("Mar")
                elif y == 2026 and m == 4:
                    months.append("Apr")
                elif y == 2026 and m == 5:
                    months.append("May")
                else:
                    months.append(f"{calendar.month_abbr[m]}'{str(y)[-2:]}")
        logging.debug(f"產生連續六個月的標題: {months}")
        return months
    except Exception as e:
        logging.error(f"產生六個月標題時發生錯誤: {e}")
        return []

def find_multilevel_header(ws, main_titles, month_titles, search_limit=30):
    try:
        for r in range(1, search_limit):
            row1 = [str(cell.value).strip() if cell.value else "" for cell in ws[r]]
            main_hits = [h for h in main_titles if h in row1]
            if len(main_hits) >= 2:
                row2 = [str(cell.value).strip() if cell.value else "" for cell in ws[r+1]]
                month_hits = [m for m in month_titles if m in row2]
                if len(month_hits) >= 2:
                    logging.info(f"找到多層標題列於 row {r} 和 row {r+1} (主標題: {main_hits}, 月份: {month_hits})")
                    return r, r+1
        logging.warning("find_multilevel_header 未找到符合條件的雙列標題")
        raise ValueError("找不到雙列標題")
    except Exception as e:
        logging.error(f"find_multilevel_header 發生錯誤: {e}")
        raise

def merge_multilevel_header(header1, header2):
    try:
        header1_filled = header1.copy()
        last_main = ""
        for i in range(len(header1_filled)):
            if header1_filled[i]:
                last_main = header1_filled[i]
            else:
                header1_filled[i] = last_main

        final_headers = []
        for h1, h2 in zip(header1_filled, header2):
            h1 = h1.strip() if isinstance(h1, str) else ""
            h2 = h2.strip() if isinstance(h2, str) else ""
            if h1 and h2:
                final_headers.append(f"{h1}_{h2}")
            elif not h1 and h2:
                final_headers.append(h2)
            elif h1 and not h2:
                final_headers.append(h1)
            else:
                final_headers.append("")
        logging.debug(f"merge_multilevel_header 輸出: {final_headers}")
        return final_headers
    except Exception as e:
        logging.error(f"merge_multilevel_header 發生錯誤: {e}")
        return []

def make_headers_unique(headers):
    try:
        counts = {}
        result = []
        for h in headers:
            if h == "":
                result.append("")
                continue
            if h not in counts:
                counts[h] = 1
                result.append(h)
            else:
                counts[h] += 1
                result.append(f"{h}_{counts[h]}")
        if len(result) != len(set(result)):
            logging.warning(f"make_headers_unique 結果還有重複標題：{result}")
        else:
            logging.debug(f"make_headers_unique 處理結果: {result}")
        return result
    except Exception as e:
        logging.error(f"make_headers_unique 發生錯誤: {e}（headers={headers}）")
        return []

def extract_group_month_fields(headers):
    try:
        group_fields = {}
        pattern = r"^(.+)_((?:Jan|Feb|Mar|Apr|May|Jun'25|Jun|Jul|Aug|Sep|Oct|Nov|Dec)(?:'\d\d|\-\d\d\d\d)?)(?:_(\d+))?$"
        for h in headers:
            try:
                m = re.match(pattern, h)
                if m:
                    group = m.group(1)
                    month = m.group(2)
                    if group not in group_fields:
                        group_fields[group] = []
                    group_fields[group].append((month, h))
            except Exception as single_err:
                logging.warning(f"extract_group_month_fields 單一標題 {h} 匹配時出錯: {single_err}")
        logging.debug(f"extract_group_month_fields 分組結果: {group_fields}")
        return group_fields
    except Exception as e:
        logging.error(f"extract_group_month_fields 發生嚴重錯誤: {e}（headers={headers}）")
        return {}

def next_n_month_names(n, header_months):
    try:
        now = datetime.now()
        y, m = now.year, now.month
        months = []
        cnt = 0
        all_header_months = set(header_months)
        while cnt < n:
            this_year = y
            m_str_std = datetime(this_year, m, 1).strftime("%b")
            m_str_jun25 = "Jun'25" if m == 6 and "Jun'25" in all_header_months else None
            m_str_year = f"{this_year}-{m_str_std}" if f"{this_year}-{m_str_std}" in all_header_months else None
            if m_str_jun25:
                m_str = m_str_jun25
            elif m_str_year:
                m_str = m_str_year
            else:
                m_str = m_str_std
            months.append(m_str)
            m += 1
            if m > 12:
                m = 1
                y += 1
            cnt += 1
        logging.debug(f"next_n_month_names 計算結果: {months}（header_months={header_months}）")
        return months
    except Exception as e:
        logging.error(f"next_n_month_names 發生錯誤: {e}（n={n}, header_months={header_months}）")
        return []

def draw_table(ws, start_row):
    try:
        # 設定欄寬
        for i, width in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width

        # 合併儲存格
        try:
            ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row+1, end_column=1)
            ws.merge_cells(start_row=start_row, start_column=2, end_row=start_row+1, end_column=2)
            ws.merge_cells(start_row=start_row, start_column=3, end_row=start_row+1, end_column=3)
            ws.merge_cells(start_row=start_row, start_column=4, end_row=start_row+1, end_column=4)
            ws.merge_cells(start_row=start_row, start_column=5, end_row=start_row, end_column=10)
            ws.merge_cells(start_row=start_row, start_column=11, end_row=start_row+1, end_column=11)
            ws.merge_cells(start_row=start_row+2, start_column=1, end_row=start_row+4, end_column=1)
            ws.merge_cells(start_row=start_row+2, start_column=2, end_row=start_row+4, end_column=2)
            ws.merge_cells(start_row=start_row+2, start_column=3, end_row=start_row+4, end_column=3)
            ws.merge_cells(start_row=start_row+2, start_column=11, end_row=start_row+4, end_column=11)
        except Exception as e_merge:
            logging.error(f"draw_table 合併儲存格時發生錯誤，起始列 {start_row}，錯誤訊息：{e_merge}")

        #  標題內容與格式
        try:
            ws.cell(row=start_row, column=1, value="Process")
            ws.cell(row=start_row, column=2, value="Tester")
            ws.cell(row=start_row, column=3, value="Customer")
            ws.cell(row=start_row, column=4, value="Month")
            ws.cell(row=start_row, column=5, value="6M FCST ( pcs/wk )")
            ws.cell(row=start_row, column=11, value="Summary")

            for col in range(1, 12):
                cell = ws.cell(row=start_row, column=col)
                cell.fill = header_fill
                cell.font = calibri_bold
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = border

            for col in range(5, 11):
                cell = ws.cell(row=start_row+1, column=col)
                cell.fill = header_fill
                cell.font = calibri_bold
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = border

            for col in range(1, 12):
                cell = ws.cell(row=start_row+1, column=col)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = border
        except Exception as e_title:
            logging.warning(f"draw_table 設定標題失敗，起始列 {start_row}，錯誤訊息：{e_title}")
        
        # 直排內容與格式
        try:
            ws.cell(row=start_row+2, column=4, value="MachineO/H")
            ws.cell(row=start_row+3, column=4, value="Machine require(set/Wk)")
            ws.cell(row=start_row+4, column=4, value="idling tester")
            for row in range(start_row+2, start_row+5):
                cell = ws.cell(row=row, column=4)
                cell.fill = header_fill
                cell.font = calibri_bold
                cell.alignment = Alignment(horizontal="left", vertical="center")
                cell.border = border
        except Exception as e_row:
            logging.warning(f"draw_table 設定直排內容失敗，起始列 {start_row}，錯誤訊息：{e_row}")

        # 補其餘區塊邊框
        try:
            for r in range(start_row, start_row+5):
                for c in range(1, 12):
                    ws.cell(row=r, column=c).border = border
        except Exception as e_border:
            logging.warning(f"draw_table 設定邊框失敗，起始列 {start_row}，錯誤訊息：{e_border}")

    except Exception as e:
        print(f"draw_table 發生不可預期錯誤，起始列 {start_row}，錯誤訊息：{e}")

def write_table_block(ws, start_row, data, months, process, tester, customer):
    for i, row_offset in enumerate([2, 3, 4]):
        r = start_row + row_offset
        if row_offset == 2:
            for col, val, name in zip([1, 2, 3], [process, tester, customer], ["Process", "Tester", "Customer"]):
                try:
                    cell = ws.cell(row=r, column=col, value=val)
                    cell.font = calibri_bold
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                except Exception as e:
                    logging.warning(f"write_table_block 寫入 {name}({val}) 發生錯誤, row={r}, col={col}：{e}")
        
        for j, m in enumerate(months):
            try:
                if i == 0:
                    cell_month = ws.cell(row=r-1, column=5+j, value=m)
                    cell_month.font = calibri_bold
                    cell_month.alignment = Alignment(horizontal="center", vertical="center")
            except Exception as e:
                logging.warning(f"write_table_block 寫入 Month 標題({m}) 發生錯誤, row={r-1}, col={5+j}：{e}")

            try:    
                if i == 0:
                    value = data['Machine O/H'][j]
                elif i == 1:
                    value = data['Machine require(set/Wk)'][j]
                elif i == 2:
                    value = data['idling tester'][j]
                else:
                    value = None
                
                cell = ws.cell(row=r, column=5+j, value=value)
                cell.font = calibri_bold
                cell.alignment = Alignment(horizontal="center", vertical="center")
                if isinstance(value, (int, float)):
                    cell.number_format = '0.0'
                if i == 2 and value is not None:
                    try:
                        if float(value) < 0:
                            cell.fill = red_fill
                    except Exception:
                        logging.warning(f"write_table_block 判斷負值時出錯, value={value}, row={r}, col={5+j}：{ve}")
            except Exception as e:
                logging.warning(f"write_table_block 寫入資料 ({value}) 時出錯, row={r}, col={5+j}：{e}")

def process_cp_sheet(wb, ws_name, output_ws, start_row, main_titles):
    logger = logging.getLogger("excel_proceed.process_cp_sheet")
    try:
        ws = wb[ws_name]
        main_titles = main_titles
        months = gen_next_6months_titles()
        
        header_row1, header_row2 = find_multilevel_header(ws, main_titles + ["Process", "Tester", "Customer"], months)
        header1 = [str(cell.value).strip() if cell.value else "" for cell in ws[header_row1]]
        header2 = [str(cell.value).strip() if cell.value else "" for cell in ws[header_row2]]

        final_headers = merge_multilevel_header(header1, header2)
        unique_headers = make_headers_unique(final_headers)

        group_month_fields = extract_group_month_fields(unique_headers)
        all_months = sorted({month for fields in group_month_fields.values() for (month, _) in fields})

        target_months = next_n_month_names(6, all_months)
        logger.info(f"{ws_name}: 自動判斷本月起六個月：{target_months}")

        fill_columns = ["Process", "Tester", "Customer"]
        last_values = {col: None for col in fill_columns}
        skip_process = {"T1 CP Subtotal", "Utilization"}

        data_rows = []
        for row in ws.iter_rows(min_row=header_row2+1, values_only=True):
            if all(cell is None or cell == "" for cell in row):
                continue
            row_data = list(row)
            if len(row_data) < len(unique_headers):
                row_data += [None] * (len(unique_headers) - len(row_data))
            elif len(row_data) > len(unique_headers):
                row_data = row_data[:len(unique_headers)]
            row_dict = dict(zip(unique_headers, row_data))
            for col in fill_columns:
                if not row_dict.get(col):
                    row_dict[col] = last_values[col]
                else:
                    last_values[col] = row_dict[col]
            if row_dict.get('Process') in skip_process:
                continue  
            
            is_negative = False
            for m in target_months:
                val = row_dict.get(f"Idling tester (set)_{m}")
                try:
                    if val is not None and float(val) < 0:
                        is_negative = True
                        break
                except Exception as e:
                    logger.warning(f"檢查負值時轉換失敗: Process={row_dict.get('Process')} Tester={row_dict.get('Tester')}, Month={m}, Value={val}, 錯誤:{e}")
            if not is_negative:
                continue
            data_rows.append(row_dict)

        grouped = defaultdict(list)
        for row in data_rows:
            key = (row['Process'], row['Tester'], row['Customer'])
            grouped[key].append(row)

        for (process, tester, customer), rows in grouped.items():
            row = rows[0]
            data = {
                'Machine O/H': [row.get(f'Machine O/H (set)_{m}') for m in target_months],
                'Machine require(set/Wk)': [row.get(f'Machine require(set/Wk)_{m}') for m in target_months],
                'idling tester': [row.get(f'Idling tester (set)_{m}') for m in target_months],
            }
            try:
                draw_table(output_ws, start_row)
                write_table_block(output_ws, start_row, data, target_months, process, tester, customer)
            except Exception as e:
                logger.error(f"寫表格失敗: Process={process}, Tester={tester}, Customer={customer}, 錯誤: {e}")
            start_row += 6  # 每個表格區塊往下推6列（含標題、資料3列、1列空行）
        logger.info(f"{ws_name} 完成, 共寫入 {len(grouped)} 個表格區塊")
        return start_row
    except Exception as e:
        logger.error(f"{ws_name} 處理發生重大錯誤: {e}", exc_info=True)
        raise

def split_ft_tables(ws, search_limit=100):
    """
    回傳: (第一表格起始列, 第一表格結束列, 第二表格起始列, 第二表格結束列)
    """
    logger = logging.getLogger("excel_proceed.split_ft_tables")
    logger.info("開始尋找 FT Summary 兩個表格的區間（search_limit=%s）", search_limit)

    try:
        empty_row_count = 0
        first_table_start = None
        first_table_end = None
        second_table_start = None
        second_table_end = None

        for idx, row in enumerate(ws.iter_rows(min_row=1, max_row=search_limit, values_only=True), start=1):
            for cell in row:
                if cell == "Process":
                    if first_table_start is None:
                        first_table_start = idx - 1
                        logger.debug(f"發現上表 Process，first_table_start = {first_table_start}")

                elif cell == "Major":
                    if second_table_start is None:
                        second_table_start = idx
                        first_table_end = idx - 3
                        logger.debug(f"發現下表 Major，second_table_start = {second_table_start}, first_table_end = {first_table_end}")

            # 檢查是否為空行
            if all(cell is None or cell == "" for cell in row[:10]):
                empty_row_count += 1
                if empty_row_count >= 2:
                    if first_table_start is not None and first_table_end is not None and second_table_start is not None:
                        second_table_end = idx - 1
                        logger.info(f"區間判定完成，上表: {first_table_start}-{first_table_end}, 下表: {second_table_start}-{second_table_end}")                     
                        break

        if first_table_end is None or second_table_start is None:
            logger.error("未找到兩個表格的分隔點，請確認 FT Summary 格式是否正確！")
            raise ValueError("未找到兩個表格的分隔點，請確認 FT Summary 格式是否正確！")

        return (first_table_start, first_table_end, second_table_start, second_table_end)
    
    except Exception as e:
        logger.exception(f"split_ft_tables 發生錯誤：{e}")
        raise

def parse_table(ws, start_row, end_row):
    logger = logging.getLogger("excel_proceed.parse_table")
    logger.info(f"開始解析表格區間 rows {start_row}-{end_row}")
    try:
        header_row1 = [str(cell.value).strip() if cell.value else "" for cell in ws[start_row]]
        header_row2 = [str(cell.value).strip() if cell.value else "" for cell in ws[start_row + 1]]
        headers = merge_multilevel_header(header_row1, header_row2)
        headers = make_headers_unique(headers)
        logger.debug(f"headers = {headers}")

        process_data = ""
        # 先找第一個有效的 process，當作初始值
        for row in ws.iter_rows(min_row=start_row + 2, max_row=end_row, values_only=True):
            if all(cell is None or cell == "" for cell in row):
                continue
            if row[0] == "" or row[0] is None:
                continue
            else:
                process_data = row[0]
                logger.debug(f"第一筆 Process: {process_data}")
                break
        
        data_rows = []
        last_vals = [None, None, None, None]
        last_vals[0] = process_data  
        for row in ws.iter_rows(min_row=start_row + 2, max_row=end_row, values_only=True):
            if all(cell is None or cell == "" for cell in row):
                continue
            row = list(row)
            if row[0] == "CP sub total":
                continue
            # 補齊空欄
            while len(row) < len(headers):
                row.append(None)
            
            for i in range(4):  # Customer 欄位在第 4 列，第三列為空
                if i == 2:
                    continue
                if row[i]:
                    last_vals[i] = row[i]
                else:
                    if i == 3:
                        # Customer 欄位如果沒有值，則不補上最後值
                        continue
                    row[i] = last_vals[i]
            
            row_dict = {}
            for idx, h in enumerate(headers):
                if h.startswith("Machine Balance"):
                    h = h.replace("Machine Balance", "Idling tester")
                row_dict[h] = row[idx] if idx < len(row) else None

            row_dict["Process"] = row[0]
            row_dict["Tester"] = row[1]
            row_dict["Customer"] = row[3]
            data_rows.append(row_dict)
        logger.info(f"成功解析表格資料，共 {len(data_rows)} 筆")
        return data_rows, headers
    
    except Exception as e:
        logger.exception(f"parse_table 錯誤（rows {start_row}-{end_row}）：{e}")
        raise

def merge_tables_by_key(table1, table2):
    """
    以 (Process, Tester, Customer) 為 key 合併上下表資料，欄位重疊以下表為主
    """
    logger = logging.getLogger("excel_proceed.merge_tables_by_key")
    logger.info(f"開始合併 table1({len(table1)}) + table2({len(table2)})")

    try:
        merged = defaultdict(dict)
        for row in table1:
            key = (row["Process"], row["Tester"], row["Customer"])
            merged[key].update(row)
        for row in table2:
            key = (row["Process"], row["Tester"], row["Customer"])
            merged[key].update(row)  # 下表補上欄位，如已存在以下表為主
        merged_list = [v for v in merged.values()]
        logger.info(f"合併完成，共 {len(merged_list)} 組 keys")
        return merged_list
    except Exception as e:
        logger.exception(f"merge_tables_by_key 錯誤：{e}")
        raise       

def process_ft_sheet(wb, ws_name, output_ws, start_row, main_titles):
    logger = logging.getLogger("excel_proceed.process_ft_sheet")

    try:
        written_tables = 0
        logger.info(f"處理工作表: {ws_name}")
        ws = wb[ws_name]
        main_titles = main_titles
        months = gen_next_6months_titles()
        logger.debug(f"產生月份名稱: {months}")

        ft1_start, ft1_end, ft2_start, ft2_end = split_ft_tables(ws)
        logger.info(f"FT Summary 兩表範圍: ({ft1_start}, {ft1_end}), ({ft2_start}, {ft2_end})")

        table1, headers1 = parse_table(ws, ft1_start, ft1_end)
        table2, headers2 = parse_table(ws, ft2_start, ft2_end)
        logger.info(f"上表資料筆數: {len(table1)}，下表資料筆數: {len(table2)}")

        merged_data = merge_tables_by_key(table1, table2)
        logger.info(f"合併後資料組數: {len(merged_data)}")

        ws_target = output_ws

        all_months = []
        for row in merged_data:
            for k in row.keys():
                if re.match(r".+_(?:Jan|Feb|Mar|Apr|May|Jun'25|Jun|Jul|Aug|Sep|Oct|Nov|Dec)(?:'\d\d|\-\d\d\d\d)?$", k):
                    m = k.split("_")[-1]
                    if m not in all_months:
                        all_months.append(m)
        target_months = next_n_month_names(6, all_months)
        logger.info(f"自動取得六個月: {target_months}")

        grouped = defaultdict(list)
        for row in merged_data:
            key = (row['Process'], row['Tester'], row['Customer'])
            grouped[key].append(row)

        for (process, tester, customer), rows in grouped.items():
            row = rows[0] 
            data = {
                'Machine O/H': [row.get(f"Machine O/H (set)_{m}") for m in target_months],
                'Machine require(set/Wk)': [row.get(f"Machine require(set/Wk)_{m}") for m in target_months],
                'idling tester': [row.get(f"Idling tester (set)_{m}") for m in target_months],
            }
            # 判斷是否有任一 idling tester 為負才畫表格
            try:
                negative = any(
                    x is not None
                    and isinstance(x, (int, float, str))
                    and str(x).replace('.', '', 1).replace('-', '', 1).isdigit()
                    and float(x) < 0
                    for x in data['idling tester']
                )
            except Exception as ex:
                logger.warning(f"比對 idling tester 負值時出錯: ({process}, {tester}, {customer})，錯誤：{ex}")
                negative = False
            
            if negative:
                try:
                    draw_table(ws_target, start_row)
                    write_table_block(ws_target, start_row, data, target_months, process, tester, customer)
                    written_tables += 1
                except Exception as e:
                    logger.error(f"寫入表格失敗: {process}, {tester}, {customer}，錯誤：{e}")
                start_row += 6
        logger.info(f"完成寫入 {written_tables} 組表格")
        return start_row
    except Exception as e:
        logger.exception(f"process_ft_sheet 執行失敗: {e}")
        raise

def prepare_environment():
    try:
        for folder in [RESOURCE_DIR, EXPORT_DIR, LOG_DIR]:
            if not os.path.exists(folder):
                os.makedirs(folder)
        # 清空 export
        for f in os.listdir(EXPORT_DIR):
            file_path = os.path.join(EXPORT_DIR, f)
            if os.path.isfile(file_path):
                os.remove(file_path)
        # 設定 log
        today_str = datetime.now().strftime('%Y%m%d')
        log_file = os.path.join(LOG_DIR, f'exec_{today_str}.log')
        logging.basicConfig(
            filename=log_file,
            filemode='a',
            format='%(asctime)s [%(levelname)s] %(message)s',
            level=logging.INFO,
            encoding='utf-8'
        )
        logging.info("=== 程式開始 ===")
    except Exception as e:
        print(f"環境準備失敗: {e}")
        logging.error(f"環境準備失敗: {e}")
        raise

def get_excel_file_list():
    files = [f for f in os.listdir(RESOURCE_DIR) if f.endswith('.xlsm')]
    if not files:
        logging.error("resource 資料夾找不到 xlsm 檔案")
        raise FileNotFoundError("resource 資料夾找不到 xlsm 檔案")
    return files

def get_or_create_clear_sheet(wb, sheet_name):
    # 如果已存在，清空內容；不存在就建立一個
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)
    return ws

def process_file(source_path, export_path):
    wb = load_workbook(source_path, keep_vba=True)

    # 建立或清空 Data Presentation
    ws_target = get_or_create_clear_sheet(wb, "Data presentation")
    
    start_row = 1
    ws_name = "CP Summary"
    main_titles = ['Machine O/H (set)', 'Machine require(set/Wk)', 'Idling tester (set)']

    logging.info(f"開始處理 {ws_name}")
    start_row = process_cp_sheet(wb, ws_name, ws_target, start_row, main_titles)
    logging.info(f"{ws_name} 完成，已畫出 CP Summary 表格共： {(start_row-1)//6} 個區塊")

    ws_name = "FT Summary"
    logging.info(f"開始處理 {ws_name}")
    start_row = process_ft_sheet(wb, ws_name, ws_target, start_row, main_titles)
    logging.info(f"{ws_name} 完成，已畫出 FT Summary 表格共： {(start_row-1)//6} 個區塊(累計)")

    wb.save(export_path)
    logging.info(f"已儲存輸出檔案: {export_path}")

RESOURCE_DIR = 'resource'
EXPORT_DIR = 'export'
LOG_DIR = 'log'

if __name__ == "__main__":
    try:
        prepare_environment()
        excel_files = get_excel_file_list()
        if not excel_files:
            raise FileNotFoundError("在 resource 資料夾中找不到任何 xlsm 檔案。")

        for file in excel_files:
            logging.info(f"處理檔案: {file}")
            source_path = os.path.join(RESOURCE_DIR, file)
            export_path = os.path.join(EXPORT_DIR, f"{os.path.splitext(file)[0]}_export.xlsm")

            try:
                process_file(source_path, export_path)
                logging.info(f"檔案 {file} 處理成功，輸出到 {export_path}")
            except Exception as e:
                logging.error(f"處理檔案 {file} 時發生錯誤: {e}")
                print(f"處理檔案 {file} 時發生錯誤: {e}")

        print("所有表格處理與輸出成功完成。")
    except Exception as e:
        logging.exception(f"主流程發生未預期錯誤: {e}")
        print(f"主流程發生未預期錯誤: {e}")