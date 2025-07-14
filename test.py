from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import calendar
import re
from collections import defaultdict

# 欄位寬度與格式設定
col_widths = [16, 22, 22, 25, 12, 12, 12, 12, 12, 12, 14]
header_fill = PatternFill(start_color="FDE8D7", end_color="FDE8D7", fill_type="solid")
bold_font = Font(bold=True)
calibri_bold = Font(name='Calibri', bold=True)
thin = Side(border_style="thin", color="000000")
red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
border = Border(top=thin, left=thin, right=thin, bottom=thin)

def gen_next_6months_titles():
    months = []
    now = datetime.now()
    for i in range(6):
        y = now.year
        m = now.month + i
        if m > 12:
            y += (m - 1) // 12
            m = (m - 1) % 12 + 1
        if i == 0 and now.month == 6 and now.year == 2025:
            months.append("Jun'25")
        else:
            # 根據資料表格實際格式決定月份命名
            if y == 2025 and m == 6:
                months.append("Jun'25")
            elif y == 2025 and m == 7:
                months.append("Jul")
            elif y == 2025 and m == 8:
                months.append("Aug")
            elif y == 2025 and m == 9:
                months.append("Sep")
            elif y == 2025 and m == 10:
                months.append("Oct")
            elif y == 2025 and m == 11:
                months.append("Nov")
            elif y == 2025 and m == 12:
                months.append("Dec")
            elif y == 2026 and m == 1:
                months.append("2026-Jan")
            elif y == 2026 and m == 2:
                months.append("Feb")
            elif y == 2026 and m == 3:
                months.append("Mar")
            elif y == 2026 and m == 4:
                months.append("Apr")
            elif y == 2026 and m == 5:
                months.append("May")
            else:
                months.append(f"{calendar.month_abbr[m]}'{str(y)[-2:]}")
    return months

def find_multilevel_header(ws, main_titles, month_titles, search_limit=30):
    for r in range(1, search_limit):
        row1 = [str(cell.value).strip() if cell.value else "" for cell in ws[r]]
        main_hits = [h for h in main_titles if h in row1]
        if len(main_hits) >= 2:
            row2 = [str(cell.value).strip() if cell.value else "" for cell in ws[r+1]]
            month_hits = [m for m in month_titles if m in row2]
            if len(month_hits) >= 2:
                return r, r+1
    raise ValueError("找不到雙列標題")

def merge_multilevel_header(header1, header2):
    header1_filled = header1.copy()
    last_main = ""
    for i in range(len(header1_filled)):
        if header1_filled[i]:
            last_main = header1_filled[i]
        else:
            header1_filled[i] = last_main

    final_headers = []
    for h1, h2 in zip(header1_filled, header2):
        h1 = h1.strip() if isinstance(h1, str) else ""
        h2 = h2.strip() if isinstance(h2, str) else ""
        if h1 and h2:
            final_headers.append(f"{h1}_{h2}")
        elif not h1 and h2:
            final_headers.append(h2)
        elif h1 and not h2:
            final_headers.append(h1)
        else:
            final_headers.append("")
    return final_headers

def make_headers_unique(headers):
    counts = {}
    result = []
    for h in headers:
        if h == "":
            result.append("")
            continue
        if h not in counts:
            counts[h] = 1
            result.append(h)
        else:
            counts[h] += 1
            result.append(f"{h}_{counts[h]}")
    return result

def extract_group_month_fields(headers):
    group_fields = {}
    pattern = r"^(.+)_((?:Jan|Feb|Mar|Apr|May|Jun'25|Jun|Jul|Aug|Sep|Oct|Nov|Dec)(?:'\d\d|\-\d\d\d\d)?)(?:_(\d+))?$"
    for h in headers:
        m = re.match(pattern, h)
        if m:
            group = m.group(1)
            month = m.group(2)
            if group not in group_fields:
                group_fields[group] = []
            group_fields[group].append((month, h))
    return group_fields

def next_n_month_names(n, header_months):
    now = datetime.now()
    y, m = now.year, now.month
    months = []
    cnt = 0
    all_header_months = set(header_months)
    while cnt < n:
        this_year = y
        m_str_std = datetime(this_year, m, 1).strftime("%b")
        m_str_jun25 = "Jun'25" if m == 6 and "Jun'25" in all_header_months else None
        m_str_year = f"{this_year}-{m_str_std}" if f"{this_year}-{m_str_std}" in all_header_months else None
        if m_str_jun25:
            m_str = m_str_jun25
        elif m_str_year:
            m_str = m_str_year
        else:
            m_str = m_str_std
        months.append(m_str)
        m += 1
        if m > 12:
            m = 1
            y += 1
        cnt += 1
    return months

def draw_table(ws, start_row):
    try:
        # 設定欄寬
        for i, width in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width

        # 合併儲存格
        ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row+1, end_column=1)
        ws.merge_cells(start_row=start_row, start_column=2, end_row=start_row+1, end_column=2)
        ws.merge_cells(start_row=start_row, start_column=3, end_row=start_row+1, end_column=3)
        ws.merge_cells(start_row=start_row, start_column=4, end_row=start_row+1, end_column=4)
        ws.merge_cells(start_row=start_row, start_column=5, end_row=start_row, end_column=10)
        ws.merge_cells(start_row=start_row, start_column=11, end_row=start_row+1, end_column=11)
        ws.merge_cells(start_row=start_row+2, start_column=1, end_row=start_row+4, end_column=1)
        ws.merge_cells(start_row=start_row+2, start_column=2, end_row=start_row+4, end_column=2)
        ws.merge_cells(start_row=start_row+2, start_column=3, end_row=start_row+4, end_column=3)
        ws.merge_cells(start_row=start_row+2, start_column=11, end_row=start_row+4, end_column=11)

        # 標題內容與格式
        ws.cell(row=start_row, column=1, value="Process")
        ws.cell(row=start_row, column=2, value="Tester")
        ws.cell(row=start_row, column=3, value="Customer")
        ws.cell(row=start_row, column=4, value="Month")
        ws.cell(row=start_row, column=5, value="6M FCST ( pcs/wk )")
        ws.cell(row=start_row, column=11, value="Summary")

        for col in range(1, 12):
            cell = ws.cell(row=start_row, column=col)
            cell.fill = header_fill
            cell.font = calibri_bold
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border

        for col in range(5, 11):
            cell = ws.cell(row=start_row+1, column=col)
            cell.fill = header_fill
            cell.font = calibri_bold
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border

        for col in range(1, 12):
            cell = ws.cell(row=start_row+1, column=col)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border

        # 直排內容與格式
        ws.cell(row=start_row+2, column=4, value="MachineO/H")
        ws.cell(row=start_row+3, column=4, value="Machine require(set/Wk)")
        ws.cell(row=start_row+4, column=4, value="idling tester")
        for row in range(start_row+2, start_row+5):
            cell = ws.cell(row=row, column=4)
            cell.fill = header_fill
            cell.font = calibri_bold
            cell.alignment = Alignment(horizontal="left", vertical="center")
            cell.border = border

        # 補其餘區塊邊框
        for r in range(start_row, start_row+5):
            for c in range(1, 12):
                ws.cell(row=r, column=c).border = border

    except Exception as e:
        print(f"draw_table 發生不可預期錯誤，起始列 {start_row}，錯誤訊息：{e}")

def write_table_block(ws, start_row, data, months, process, tester, customer):
    for i, row_offset in enumerate([2, 3, 4]):
        r = start_row + row_offset
        # 左側 Process, Tester, Customer 只寫第一列
        if row_offset == 2:
            for col, val in zip([1, 2, 3], [process, tester, customer]):
                cell = ws.cell(row=r, column=col, value=val)
                cell.font = calibri_bold
                cell.alignment = Alignment(horizontal="center", vertical="center")
        # 六個月橫向資料
        for j, m in enumerate(months):
            # 最上方那一行填 Month（row=r-1，只在 Machine O/H 那一列填一次）
            if i == 0:
                cell_month = ws.cell(row=r-1, column=5+j, value=m)
                cell_month.font = calibri_bold
                cell_month.alignment = Alignment(horizontal="center", vertical="center")
            # 實際數據格填資料
            if i == 0:
                value = data['Machine O/H'][j]
            elif i == 1:
                value = data['Machine require(set/Wk)'][j]
            elif i == 2:
                value = data['idling tester'][j]
            cell = ws.cell(row=r, column=5+j, value=value)
            cell.font = calibri_bold
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if isinstance(value, (int, float)):
                cell.number_format = '0.0'
            if i == 2 and value is not None:
                try:
                    if float(value) < 0:
                        cell.fill = red_fill
                except Exception:
                    pass

def process_cp_sheet(wb, ws_name, output_ws, start_row, main_titles):
    ws = wb[ws_name]
    # 你的設定
    main_titles = main_titles
    months = gen_next_6months_titles()
    
    header_row1, header_row2 = find_multilevel_header(ws, main_titles + ["Process", "Tester", "Customer"], months)
    header1 = [str(cell.value).strip() if cell.value else "" for cell in ws[header_row1]]
    header2 = [str(cell.value).strip() if cell.value else "" for cell in ws[header_row2]]

    # 組合欄位名稱
    final_headers = merge_multilevel_header(header1, header2)
    unique_headers = make_headers_unique(final_headers)

    # 取得所有主題_月份型態的欄位
    group_month_fields = extract_group_month_fields(unique_headers)
    all_months = sorted({month for fields in group_month_fields.values() for (month, _) in fields})

    # 取得當月起六個月的月份名（以所有 header 出現過的月份為依據）
    target_months = next_n_month_names(6, all_months)
    print("自動判斷本月起六個月：", target_months)

    fill_columns = ["Process", "Tester", "Customer"]
    last_values = {col: None for col in fill_columns}
    skip_process = {"T1 CP Subtotal", "Utilization"}

    # 讀取資料列並補齊合併欄
    data_rows = []
    for row in ws.iter_rows(min_row=header_row2+1, values_only=True):
        if all(cell is None or cell == "" for cell in row):
            continue
        row_data = list(row)
        if len(row_data) < len(unique_headers):
            row_data += [None] * (len(unique_headers) - len(row_data))
        elif len(row_data) > len(unique_headers):
            row_data = row_data[:len(unique_headers)]
        row_dict = dict(zip(unique_headers, row_data))
        for col in fill_columns:
            if not row_dict.get(col):
                row_dict[col] = last_values[col]
            else:
                last_values[col] = row_dict[col]
        if row_dict.get('Process') in skip_process:
            continue  
        
        is_negative = False
        for m in target_months:
            val = row_dict.get(f"Idling tester (set)_{m}")
            try:
                if val is not None and float(val) < 0:
                    is_negative = True
                    break
            except Exception:
                continue
        if not is_negative:
            continue
        data_rows.append(row_dict)

    # 依 (Process, Tester, Customer) 分群組畫多個表
    grouped = defaultdict(list)
    for row in data_rows:
        key = (row['Process'], row['Tester'], row['Customer'])
        grouped[key].append(row)

    start_row = start_row
    for (process, tester, customer), rows in grouped.items():
        row = rows[0]
        data = {
            'Machine O/H': [row.get(f'Machine O/H (set)_{m}') for m in target_months],
            'Machine require(set/Wk)': [row.get(f'Machine require(set/Wk)_{m}') for m in target_months],
            'idling tester': [row.get(f'Idling tester (set)_{m}') for m in target_months],
        }
        draw_table(ws_target, start_row)
        write_table_block(ws_target, start_row, data, target_months, process, tester, customer)
        start_row += 6  # 每個表格區塊往下推6列（含標題、資料3列、1列空行）
    return start_row

def split_ft_tables(ws, search_limit=100):
    """
    回傳: (第一表格起始列, 第一表格結束列, 第二表格起始列, 第二表格結束列)
    """
    empty_row_count = 0
    first_table_start = None
    first_table_end = None
    second_table_start = None
    second_table_end = None

    # 假設資料從第1列開始（根據你的實際格式可以調整起始列）
    for idx, row in enumerate(ws.iter_rows(min_row=1, max_row=search_limit, values_only=True), start=1):
        for cell in row:
            if cell == "Process":
                if first_table_start is None:
                    first_table_start = idx - 1
            elif cell == "Major":
                if second_table_start is None:
                    second_table_start = idx
                    first_table_end = idx - 3

        # 檢查是否為空行
        if all(cell is None or cell == "" for cell in row[:10]):
            empty_row_count += 1
            if empty_row_count >= 2:
                if first_table_start is not None and first_table_end is not None and second_table_start is not None:
                    second_table_end = idx - 1
                    break

    # 如果沒有連續兩個空行，只能設計為單表格處理
    if first_table_end is None or second_table_start is None:
        raise ValueError("未找到兩個表格的分隔點，請確認 FT Summary 格式是否正確！")

    return (first_table_start, first_table_end, second_table_start, second_table_end)

def parse_table(ws, start_row, end_row):
    # 取得主副標題
    header_row1 = [str(cell.value).strip() if cell.value else "" for cell in ws[start_row]]
    header_row2 = [str(cell.value).strip() if cell.value else "" for cell in ws[start_row + 1]]
    headers = merge_multilevel_header(header_row1, header_row2)
    headers = make_headers_unique(headers)

    # 產生欄位名稱 -> 欄位索引的對應表
    col_idx = {name: idx for idx, name in enumerate(headers)}

    process_data = ""
    for row in ws.iter_rows(min_row=start_row + 2, max_row=end_row, values_only=True):
        print(row[:4])
        if all(cell is None or cell == "" for cell in row):
            continue
        if row[0] == "" or row[0] is None:
            continue
        else:
            process_data = row[0]
            break
    # 資料
    data_rows = []
    last_vals = [None, None, None, None]
    last_vals[0] = process_data  # 初始 Process 資料
    for row in ws.iter_rows(min_row=start_row + 2, max_row=end_row, values_only=True):
        if all(cell is None or cell == "" for cell in row):
            continue
        row = list(row)
        if row[0] == "CP sub total":
            continue
        # 補齊空欄
        while len(row) < len(headers):
            row.append(None)
        # 前三欄固定：Process, Tester, Customer
        for i in range(4):
            if i == 2:
                continue # 跳過空欄
            if row[i]:
                last_vals[i] = row[i]
            else:
                if i == 3:
                    # Customer 欄位如果沒有值，則不補上最後值
                    continue
                row[i] = last_vals[i]
        row_dict = {}
        for idx, h in enumerate(headers):
            # 下表格需映射欄位名稱
            if h.startswith("Machine Balance"):
                # 把 Machine Balance (set)_{月} 改成 Idling tester (set)_{月}
                h = h.replace("Machine Balance", "Idling tester")
            row_dict[h] = row[idx] if idx < len(row) else None
        # 補充：主key
        row_dict["Process"] = row[0]
        row_dict["Tester"] = row[1]
        row_dict["Customer"] = row[3]
        data_rows.append(row_dict)
    return data_rows, headers

def merge_tables_by_key(table1, table2):
    """
    以 (Process, Tester, Customer) 為 key 合併上下表資料，欄位重疊以下表為主
    """
    merged = defaultdict(dict)
    for row in table1:
        key = (row["Process"], row["Tester"], row["Customer"])
        merged[key].update(row)
    for row in table2:
        key = (row["Process"], row["Tester"], row["Customer"])
        merged[key].update(row)  # 下表補上欄位，如已存在以下表為主
    # 轉回 list of dict
    return [v for v in merged.values()]

def process_ft_sheet(wb, ws_name, output_ws, start_row, main_titles):
    ws = wb[ws_name]
    main_titles = main_titles
    months = gen_next_6months_titles()

    ft1_start, ft1_end, ft2_start, ft2_end = split_ft_tables(ws)

    table1, headers1 = parse_table(ws, ft1_start, ft1_end)
    table2, headers2 = parse_table(ws, ft2_start, ft2_end)

    merged_data = merge_tables_by_key(table1, table2)

    ws_target = output_ws

    all_months = []
    for row in merged_data:
       for k in row.keys():
            if re.match(r".+_(?:Jan|Feb|Mar|Apr|May|Jun'25|Jun|Jul|Aug|Sep|Oct|Nov|Dec)(?:'\d\d|\-\d\d\d\d)?$", k):
                # 解析出月份名
                m = k.split("_")[-1]
                if m not in all_months:
                    all_months.append(m)
    target_months = next_n_month_names(6, all_months)

    # 以 (Process, Tester, Customer) 分群組
    grouped = defaultdict(list)
    for row in merged_data:
        key = (row['Process'], row['Tester'], row['Customer'])
        grouped[key].append(row)

    start_row = 1
    for (process, tester, customer), rows in grouped.items():
        row = rows[0]  # 如一組有多筆取第一筆（可自行調整）
        data = {
            'Machine O/H': [row.get(f"Machine O/H (set)_{m}") for m in target_months],
            'Machine require(set/Wk)': [row.get(f"Machine require(set/Wk)_{m}") for m in target_months],
            'idling tester': [row.get(f"Idling tester (set)_{m}") for m in target_months],
        }
        # 判斷是否有任一 idling tester 為負才畫表格
        if any(x is not None and isinstance(x, (int, float, str)) and str(x).replace('.', '', 1).replace('-', '', 1).isdigit() and float(x) < 0 for x in data['idling tester']):
            draw_table(ws_target, start_row)
            write_table_block(ws_target, start_row, data, target_months, process, tester, customer)
            start_row += 6  # 每個表格區塊往下推6列
    return start_row

if __name__ == "__main__":
    source = "250702162359805.xlsm"
    wb = load_workbook(source, keep_vba=True)
    if "Output" in wb.sheetnames:
        del wb["Output"]
    ws_target = wb.create_sheet("Output")

    ws_name = "FT Summary"
    main_titles = ['Machine O/H (set)', 'Machine require(set/Wk)', 'Idling tester (set)']

    start_row = 1
    start_row = process_ft_sheet(wb, ws_name, ws_target, start_row, main_titles)

    output_file = f"{source.rsplit('.', 1)[0]}_test.xlsm"
    wb.save(output_file)
